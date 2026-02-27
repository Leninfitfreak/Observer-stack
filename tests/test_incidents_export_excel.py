from __future__ import annotations

from datetime import date, datetime, timezone
from io import BytesIO

from openpyxl import load_workbook
from sqlalchemy import create_engine
from sqlalchemy.orm import Session, sessionmaker

from ai_observer.backend.models.incident import Incident, IncidentMetricsSnapshot, IncidentStatusHistory
from ai_observer.backend.schemas.incidents import IncidentFilterQuery
from ai_observer.backend.services.incidents_service import IncidentsService
from ai_observer.incident_analysis.models import Base, IncidentAnalysis


def _build_service() -> tuple[IncidentsService, Session]:
    engine = create_engine("sqlite+pysqlite:///:memory:", future=True)
    Base.metadata.create_all(bind=engine)
    factory = sessionmaker(bind=engine, autoflush=False, autocommit=False, future=True)
    session: Session = factory()
    return IncidentsService(db=session), session


def test_export_excel_no_data_sheet() -> None:
    service, _session = _build_service()
    q = IncidentFilterQuery(start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), limit=20, offset=0)

    blob = service.export_excel(q)
    wb = load_workbook(filename=BytesIO(blob))

    assert wb.sheetnames == ["No data available"]
    ws = wb["No data available"]
    assert ws["A2"].value == "No data available"


def test_export_excel_full_multisheet_payload() -> None:
    service, session = _build_service()
    now = datetime.now(timezone.utc)
    incident_id = "INC-EXPORT-1"

    session.add(
        Incident(
            incident_id=incident_id,
            cluster_id="cluster-export",
            status="OPEN",
            severity="WARNING",
            impact_level="Low",
            slo_breach_risk=12.5,
            error_budget_remaining=98.0,
            affected_services="product-service,postgres",
            start_time=now,
            duration="00:05:00",
            created_at=now,
        )
    )
    session.add(
        IncidentStatusHistory(
            incident_id=incident_id,
            from_status="OPEN",
            to_status="INVESTIGATING",
            changed_at=now,
        )
    )
    session.add(
        IncidentAnalysis(
            incident_id=incident_id,
            service_name="product-service",
            cluster_id="cluster-export",
            anomaly_score=0.37,
            confidence_score=0.81,
            classification="Performance Degradation",
            root_cause="Database latency increase",
            mitigation={
                "actions": ["check db"],
                "change_detection_context": ["No deployments in last 30m"],
            },
            risk_forecast=0.29,
            mitigation_success=None,
            executive_summary="Latency increased but error rate is stable.",
            supporting_signals={"signals": ["p95 up", "db query duration up"]},
            suggested_actions={"actions": ["Inspect postgres"]},
            confidence_breakdown={"metric": 70, "trace": 55},
            created_at=now,
        )
    )
    session.add(
        IncidentMetricsSnapshot(
            incident_id=incident_id,
            cpu_usage=2.2,
            memory_usage=776.0,
            latency_p95=16.0,
            error_rate=0.0,
            thread_pool_saturation=0.12,
            raw_metrics_json={"request_rate_rps_5m": 0.28, "cpu_usage_cores_5m": 0.022},
            captured_at=now,
        )
    )
    session.commit()

    q = IncidentFilterQuery(start_date=date(2026, 1, 1), end_date=date(2026, 12, 31), limit=20, offset=0)
    blob = service.export_excel(q)
    wb = load_workbook(filename=BytesIO(blob))

    assert wb.sheetnames == ["Incident Core", "AI Analysis", "Metrics Snapshot", "Raw JSON"]

    ws_core = wb["Incident Core"]
    headers_core = [c.value for c in ws_core[1]]
    assert headers_core == [
        "Incident ID",
        "Status",
        "Severity",
        "Impact Level",
        "SLA Countdown",
        "SLO Breach Risk",
        "Error Budget Remaining",
        "Affected Services",
        "Start Time",
        "Duration",
    ]
    assert ws_core["A2"].value == incident_id

    ws_ai = wb["AI Analysis"]
    headers_ai = [c.value for c in ws_ai[1]]
    assert headers_ai == [
        "Incident ID",
        "Executive Summary",
        "Root Cause",
        "Supporting Signals",
        "Change Detection Context",
        "Risk Forecast",
        "Suggested Actions",
        "Confidence Score",
        "Confidence Breakdown",
    ]
    assert ws_ai["A2"].value == incident_id

    ws_metrics = wb["Metrics Snapshot"]
    headers_metrics = [c.value for c in ws_metrics[1]]
    assert headers_metrics == [
        "Incident ID",
        "CPU Usage",
        "Memory Usage",
        "Latency P95",
        "Error Rate",
        "Thread Pool Saturation",
        "Raw Metrics JSON",
    ]
    assert ws_metrics["A2"].value == incident_id

    ws_raw = wb["Raw JSON"]
    assert ws_raw["A1"].value == "Incident JSON"
    assert incident_id in str(ws_raw["A2"].value)
